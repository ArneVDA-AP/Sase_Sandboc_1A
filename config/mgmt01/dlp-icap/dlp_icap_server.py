#!/usr/bin/env python3
"""
DLP ICAP Server — PoC implementatie (fixed version)
Scant REQMOD POST-bodies op gevoelige data.
Draait naast ClamAV via Squid adaptation_service_chain.
"""

import re, io, os, sys, logging, socketserver, zipfile
from pyicap import ICAPServer, BaseICAPRequestHandler
from datetime import datetime, timezone
from nats_publisher import NatsPublisher


logging.basicConfig(level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)])
logger = logging.getLogger('dlp-icap')

producer = None

MAX_BODY_BYTES = 10 * 1024 * 1024

# --- Algoritmische validatiefuncties ---

def luhn_check(card_number: str) -> bool:
    digits = [int(d) for d in card_number if d.isdigit()]
    if not 13 <= len(digits) <= 19:
        return False
    checksum = 0
    for i, d in enumerate(digits[::-1]):
        if i % 2 == 1:
            d *= 2
            if d > 9: d -= 9
        checksum += d
    return checksum % 10 == 0

def iban_mod97_check(iban: str) -> bool:
    iban = iban.replace(' ', '').replace('-', '').upper()
    if not 15 <= len(iban) <= 34:
        return False
    rearranged = iban[4:] + iban[:4]
    numeric = ''.join(str(ord(c) - ord('A') + 10) if c.isalpha() else c
                      for c in rearranged if c.isalnum())
    return int(numeric) % 97 == 1

def bsn_11check(bsn: str) -> bool:
    digits = [int(d) for d in bsn if d.isdigit()]
    if len(digits) != 9:
        return False
    if digits[0] == 0:
        return False
    weights = [9, 8, 7, 6, 5, 4, 3, 2, -1]
    total = sum(d * w for d, w in zip(digits, weights))
    return total % 11 == 0

# --- Regex-patronen ---
CC_PATTERN = re.compile(r'\b([3-6]\d{3}[\s\-]?\d{4}[\s\-]?\d{4}[\s\-]?\d{3,4})\b')
IBAN_PATTERN = re.compile(r'\b([A-Z]{2}\d{2}[\s]?[A-Z0-9]{4}[\s]?\d{4}[\s]?\d{4}[\s]?\d{0,14})\b', re.I)
BSN_PATTERN = re.compile(
    r'(?:(?:bsn|burgerservicenummer|sofinummer)\s*[:\-]?\s*)(\d{9})'
    r'|(?<!\d)(\d{9})(?!\d)',
    re.I
)
CONFIDENTIAL_PATTERN = re.compile(
    r'\b(CONFIDENTIAL|VERTROUWELIJK|GEHEIM|RESTRICTED|INTERN)\b', re.I)

# --- Bestandstekst-extractie ---

def detect_file_type(data: bytes) -> str:
    if data[:5] == b'%PDF-':
        return 'pdf'
    if data[:4] == b'PK\x03\x04':
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as zf:
                names = zf.namelist()
                if any(n.startswith('word/') for n in names):
                    return 'docx'
                if any(n.startswith('xl/') for n in names):
                    return 'xlsx'
        except zipfile.BadZipFile:
            pass
        return 'zip'
    return 'text'

def extract_text(data: bytes) -> str:
    ft = detect_file_type(data)
    try:
        if ft == 'docx':
            from docx import Document
            return '\n'.join(p.text for p in Document(io.BytesIO(data)).paragraphs)
        if ft == 'xlsx':
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            return ' '.join(str(c) for ws in wb.worksheets
                           for row in ws.iter_rows(values_only=True)
                           for c in row if c is not None)
        if ft == 'pdf':
            from pypdf import PdfReader
            return '\n'.join(p.extract_text() or ''
                            for p in PdfReader(io.BytesIO(data)).pages)
    except Exception as e:
        logger.debug(f"Extractie mislukt ({ft}): {e}")
    return data.decode('utf-8', errors='ignore')

def scan_content(data: bytes) -> list:
    text = extract_text(data)
    violations = []

    for m in CC_PATTERN.finditer(text):
        clean = re.sub(r'[\s\-]', '', m.group(1))
        if luhn_check(clean):
            violations.append(('HIGH', f'Creditcard: {clean[:4]}****{clean[-4:]}'))

    for m in IBAN_PATTERN.finditer(text):
        if iban_mod97_check(m.group(1)):
            r = m.group(1).replace(' ', '')
            violations.append(('HIGH', f'IBAN: {r[:4]}****{r[-4:]}'))

    for m in BSN_PATTERN.finditer(text):
        candidate = m.group(1) or m.group(2)
        if candidate and bsn_11check(candidate):
            violations.append(('MEDIUM', f'Potentieel BSN: ***{candidate[-3:]}'))

    for m in CONFIDENTIAL_PATTERN.finditer(text):
        violations.append(('HIGH', f'Gevoeligheidslabel: {m.group(0).upper()}'))

    return violations

BLOCK_PAGE = """<!DOCTYPE html>
<html><head><title>DLP Policy Overtreding</title>
<style>
body{{font-family:-apple-system,sans-serif;background:#1a1a2e;color:#eee;
     display:flex;justify-content:center;align-items:center;min-height:100vh;margin:0}}
.card{{background:#16213e;border-left:4px solid #e94560;border-radius:8px;
       padding:2rem 3rem;max-width:600px;box-shadow:0 4px 24px rgba(0,0,0,.3)}}
h1{{color:#e94560;margin-top:0}}
.v{{background:#0f3460;padding:.8rem 1rem;border-radius:4px;margin:.5rem 0;
    font-family:monospace;font-size:.9rem}}
.HIGH{{border-left:3px solid #e94560}}.MEDIUM{{border-left:3px solid #f0a500}}
</style></head><body><div class="card">
<h1>&#x1F6E1; Upload geblokkeerd — DLP Policy</h1>
<p>Uw upload is geblokkeerd omdat gevoelige data is gedetecteerd:</p>
{violations}
<p style="color:#888;margin-top:1.5rem;font-size:.85rem">
SASE PoC — Inline DLP Engine | Neem contact op met security bij false positive.</p>
</div></body></html>"""

class DLPHandler(BaseICAPRequestHandler):
    def dlpscan_OPTIONS(self):
        self.set_icap_response(200)
        self.set_icap_header(b'Methods', b'REQMOD')
        self.set_icap_header(b'Service', b'DLP-ICAP-PoC/1.0')
        self.send_headers(False)

    def dlpscan_REQMOD(self):
        logger.debug(f"ICAP headers: {dict(self.headers)}")
        logger.debug(f"Enc req headers: {dict(self.enc_req_headers)}")
        body = b''
        size_exceeded = False

        while True:
            chunk = self.read_chunk()
            if chunk == b'':
                break
            body += chunk
            if len(body) > MAX_BODY_BYTES:
                logger.warning(
                    f"Body overschrijdt {MAX_BODY_BYTES // (1024*1024)} MB — "
                    "doorgelaten zonder DLP-scan"
                )
                while self.read_chunk() != b'':
                    pass
                size_exceeded = True
                break

        if not body or size_exceeded:
            self.no_adaptation_required()
            return

        try:
            violations = scan_content(body)
        except Exception as e:
            logger.error(f"Scan mislukt, request doorgelaten: {e}")
            self.no_adaptation_required()
            return

        if violations:
            raw_ip = self.headers.get(b'x-client-ip', [b'unknown'])[0].decode(
                errors='replace'
            )
            client_ip = re.sub(r'[\r\n\t\x00-\x1f]', '_', raw_ip)
            logger.warning(f"DLP OVERTREDING van {client_ip}: {violations}")

            if publisher is not None:
                publisher.publish("security.alert.dlp", {
                    "timestamp": datetime.now(timezone.utc).isoformat(),
                    "client_ip": client_ip,
                    "violations": [{"severity": s, "detail": d} for s, d in violations[:50]],
                    "producer": "dlp",
                })

            vhtml = '\n'.join(
                f'<div class="v {sev}">[{sev}] {desc}</div>'
                for sev, desc in violations)
            block = BLOCK_PAGE.format(violations=vhtml).encode('utf-8')

            self.set_icap_response(200)
            self.set_enc_status(b'HTTP/1.1 403 Forbidden')
            self.set_enc_header(b'Content-Type', b'text/html; charset=utf-8')
            self.set_enc_header(b'Content-Length', str(len(block)).encode())
            self.set_enc_header(b'Connection', b'close')
            self.send_headers(True)
            self.send_chunk(block)
            self.send_chunk(b'')
        else:
            self.no_adaptation_required()

class ThreadedICAPServer(socketserver.ThreadingMixIn, ICAPServer):
    allow_reuse_address = True
    max_children = 50
    daemon_threads = True

def main():
    port = int(os.environ.get('DLP_ICAP_PORT', '1345'))
    timeout = int(os.environ.get('DLP_ICAP_TIMEOUT', '30'))
    logger.info(f"DLP ICAP server gestart op 0.0.0.0:{port} (timeout={timeout}s)")
    server = ThreadedICAPServer(('0.0.0.0', port), DLPHandler)
    server.socket.settimeout(timeout)
    global publisher
    nats_url = os.environ.get('NATS_URL')
    nats_pass = os.environ.get('NATS_PASS')
    if nats_url and nats_pass:
        publisher = NatsPublisher(nats_url, os.environ.get('NATS_USER', 'mgmt01-pub'), nats_pass)
    else:
        logger.warning("NATS_URL/NATS_PASS niet gezet — DLP draait zonder NATS-publishing")
    server.serve_forever()

if __name__ == '__main__':
    main()
